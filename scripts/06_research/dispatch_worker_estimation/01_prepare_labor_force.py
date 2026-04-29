# 労働力調査 k101010.xlsx から派遣労働者割合を抽出するコード

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE_PATH = Path(__file__).resolve().parents[3] / "data"
INPUT_FILE = BASE_PATH / "00_raw" / "999_reseach" / "派遣労働者数推計" / "k101010.xlsx"
OUTPUT_DIR = BASE_PATH / "06_research" / "dispatch_worker_estimation" / "01_tidy"
OUTPUT_FILE = OUTPUT_DIR / "03_労働力調査_派遣労働者割合_tidy.csv"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(INPUT_FILE, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    employee_excluding_executives = worksheet["O23"].value
    dispatched_workers = worksheet["O29"].value
    dispatched_worker_share = dispatched_workers / employee_excluding_executives

    df = pd.DataFrame(
        [
            {
                "年": 2024,
                "対象": "全体",
                "役員を除く雇用者": employee_excluding_executives,
                "労働者派遣事業所の派遣社員": dispatched_workers,
                "派遣労働者割合": dispatched_worker_share,
                "派遣労働者割合_pct": dispatched_worker_share * 100,
                "抽出元ファイル": INPUT_FILE.name,
                "役員を除く雇用者_セル": "O23",
                "労働者派遣事業所の派遣社員_セル": "O29",
            }
        ]
    )

    df.to_csv(OUTPUT_FILE, index=False, encoding="utf-8-sig")

    print(f"役員を除く雇用者: {employee_excluding_executives}")
    print(f"労働者派遣事業所の派遣社員: {dispatched_workers}")
    print(f"派遣労働者割合: {dispatched_worker_share:.8f}")
    print(f"保存先: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
